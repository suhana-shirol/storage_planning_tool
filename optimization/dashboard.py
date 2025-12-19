import ast
import json
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

BASE_DIR = Path(__file__).resolve().parent
DECISION_VARS_CSV_PATH = BASE_DIR / "non_zero_decision_variables.csv"
PARAMS_PATH = BASE_DIR / "params.json"


@dataclass
class DashboardSections:
    raw_df: pd.DataFrame
    store_buy: Dict[str, List[Dict[str, pd.DataFrame]]]
    relocations: List[Dict[str, pd.DataFrame]]
    forward: List[Dict[str, pd.DataFrame]]
    forward_reserve: List[Dict[str, pd.DataFrame]]


def _parse_tuple_dict(raw: Dict) -> Dict[Tuple[str, str], float]:
    """Convert stringified tuple keys into real tuples for easier joins."""
    parsed: Dict[Tuple[str, str], float] = {}
    for key, val in raw.items():
        if isinstance(key, tuple) and len(key) >= 2:
            parsed[(key[0], key[1])] = val
            continue
        if isinstance(key, str) and key.startswith("("):
            try:
                tuple_key = ast.literal_eval(key)
            except Exception:
                continue
            if isinstance(tuple_key, tuple) and len(tuple_key) >= 2:
                parsed[(tuple_key[0], tuple_key[1])] = val
    return parsed


def _load_inputs(
    decision_df: Optional[pd.DataFrame] = None,
    params: Optional[Dict] = None,
    decision_vars_path: Path = DECISION_VARS_CSV_PATH,
    params_path: Path = PARAMS_PATH,
) -> Tuple[pd.DataFrame, Dict, pd.DataFrame]:
    """Load decision variables + Params, returning a normalized cStoreUnits table."""
    if decision_df is None:
        decision_df = pd.read_csv(decision_vars_path)
    else:
        decision_df = decision_df.copy()

    decision_df["Value"] = pd.to_numeric(decision_df["Value"], errors="coerce")

    if params is None:
        with open(params_path, "r") as f:
            params = json.load(f)

    cstore_pairs = _parse_tuple_dict(params.get("cStoreUnits", {}))
    cstore_rows = [(typ, loc, val) for (typ, loc), val in cstore_pairs.items()]
    if cstore_rows:
        cstore_df = pd.DataFrame(cstore_rows, columns=["Index_1", "Index_2", "cStoreUnits"])
    else:
        cstore_df = pd.DataFrame(columns=["Index_1", "Index_2", "cStoreUnits"])

    return decision_df, params, cstore_df


def _format_percent(series: pd.Series) -> pd.Series:
    """Round percentages and keep small non-zero values visible."""
    return series.replace(0, None).apply(
        lambda x: 0.01 if x is not None and 0 < x < 0.01 else round(x, 2) if x is not None else None
    )


def _build_store_buy(df: pd.DataFrame, cstore_df: pd.DataFrame) -> Dict[str, List[Dict[str, pd.DataFrame]]]:
    results: Dict[str, List[Dict[str, pd.DataFrame]]] = {"StoreUnits": [], "BuyUnits": []}

    store_df = df[df["Variable"] == "StoreUnits"]
    for loc in store_df["Index_2"].dropna().unique():
        loc_df = store_df[store_df["Index_2"] == loc]
        merged_df = loc_df.merge(cstore_df, on=["Index_1", "Index_2"], how="left")
        merged_df["cStoreUnits"] = pd.to_numeric(merged_df["cStoreUnits"], errors="coerce").fillna(0)
        merged_df["Value"] = pd.to_numeric(merged_df["Value"], errors="coerce")
        view_df = merged_df.rename(columns={"Index_1": "Storage Type", "Value": "New Storage Units"})
        view_df = view_df[["Storage Type", "New Storage Units", "cStoreUnits"]].rename(
            columns={"cStoreUnits": "Current Storage Units"}
        )
        results["StoreUnits"].append({"location": loc, "data": view_df})

    buy_df = df[df["Variable"] == "BuyUnits"]
    for loc in buy_df["Index_2"].dropna().unique():
        loc_df = buy_df[buy_df["Index_2"] == loc]
        view_df = loc_df.rename(columns={"Index_1": "Storage Type", "Value": "Units"})[["Storage Type", "Units"]]
        results["BuyUnits"].append({"location": loc, "data": view_df})

    return results


def _build_relocations(df: pd.DataFrame) -> List[Dict[str, pd.DataFrame]]:
    reloc_df = df[df["Variable"] == "RelocUnits"].copy()
    if reloc_df.empty:
        return []

    reloc_df["Index_3"] = reloc_df["Index_3"].fillna("Unknown").infer_objects(copy=False)
    tables: List[Dict[str, pd.DataFrame]] = []

    for from_loc in reloc_df["Index_2"].dropna().unique():
        df_from = reloc_df[reloc_df["Index_2"] == from_loc]
        pivot = df_from.pivot_table(
            index="Index_3", columns="Index_1", values="Value", aggfunc="sum", fill_value=0
        )
        pivot.index.name = ""
        tables.append({"from_location": from_loc, "table": pivot})

    return tables


def _build_forward(df: pd.DataFrame) -> List[Dict[str, pd.DataFrame]]:
    uni_df = df[df["Variable"] == "SKUuniquePct"].copy()
    cub_df = df[df["Variable"] == "SKUcubicPct"].copy()

    uni_df["Index_3"] = uni_df["Index_3"].fillna("Unknown")
    cub_df["Index_3"] = cub_df["Index_3"].fillna("Unknown")

    uni_df["Value"] = _format_percent(uni_df["Value"])
    cub_df["Value"] = _format_percent(cub_df["Value"])

    entries: List[Dict[str, pd.DataFrame]] = []
    for sku in sorted(uni_df["Index_1"].dropna().unique()):
        group_uni = uni_df[uni_df["Index_1"] == sku].dropna(subset=["Value"])
        pivot_uni = pd.DataFrame()
        if not group_uni.empty:
            pivot_uni = group_uni.pivot_table(
                index="Index_3", columns="Index_2", values="Value", aggfunc="sum", fill_value=0
            )
            pivot_uni.index.name = ""

        group_cub = cub_df[cub_df["Index_1"] == sku].dropna(subset=["Value"])
        pivot_cub = pd.DataFrame()
        if not group_cub.empty:
            pivot_cub = group_cub.pivot_table(
                index="Index_3", columns="Index_2", values="Value", aggfunc="sum", fill_value=0
            )
            pivot_cub.index.name = ""

        if pivot_uni.empty and pivot_cub.empty:
            continue

        entries.append({"sku": sku, "unique": pivot_uni, "cubic": pivot_cub})

    return entries


def _build_forward_reserve(df: pd.DataFrame) -> List[Dict[str, pd.DataFrame]]:
    uni_reg = df[df["Variable"] == "SKUuniquePct"].copy()
    uni_res = df[df["Variable"] == "ResSKUuniquePct"].copy()
    cub_reg = df[df["Variable"] == "SKUcubicPct"].copy()
    cub_res = df[df["Variable"] == "ResSKUcubicPct"].copy()

    for d in (uni_res, cub_res):
        if not d.empty:
            d["Index_2"] = d["Index_2"].astype(str) + "_Reserve"

    uni_df = pd.concat([uni_reg, uni_res], ignore_index=True)
    cub_df = pd.concat([cub_reg, cub_res], ignore_index=True)

    uni_df["Index_3"] = uni_df["Index_3"].fillna("Unknown")
    cub_df["Index_3"] = cub_df["Index_3"].fillna("Unknown")

    uni_df["Value"] = _format_percent(uni_df["Value"])
    cub_df["Value"] = _format_percent(cub_df["Value"])

    entries: List[Dict[str, pd.DataFrame]] = []
    for sku in sorted(uni_df["Index_1"].dropna().unique()):
        group_uni = uni_df[uni_df["Index_1"] == sku].dropna(subset=["Value"])
        pivot_uni = pd.DataFrame()
        if not group_uni.empty:
            pivot_uni = group_uni.pivot_table(
                index="Index_3", columns="Index_2", values="Value", aggfunc="sum", fill_value=0
            )
            pivot_uni.index.name = ""

        group_cub = cub_df[cub_df["Index_1"] == sku].dropna(subset=["Value"])
        pivot_cub = pd.DataFrame()
        if not group_cub.empty:
            pivot_cub = group_cub.pivot_table(
                index="Index_3", columns="Index_2", values="Value", aggfunc="sum", fill_value=0
            )
            pivot_cub.index.name = ""

        if pivot_uni.empty and pivot_cub.empty:
            continue

        entries.append({"sku": sku, "unique": pivot_uni, "cubic": pivot_cub})

    return entries


def build_dashboard_sections(
    decision_df: Optional[pd.DataFrame] = None,
    params: Optional[Dict] = None,
    decision_vars_path: Path = DECISION_VARS_CSV_PATH,
    params_path: Path = PARAMS_PATH,
) -> DashboardSections:
    df, params, cstore_df = _load_inputs(decision_df, params, decision_vars_path, params_path)
    store_buy = _build_store_buy(df, cstore_df)
    relocations = _build_relocations(df)
    forward = _build_forward(df)
    forward_reserve = _build_forward_reserve(df)
    return DashboardSections(
        raw_df=df,
        store_buy=store_buy,
        relocations=relocations,
        forward=forward,
        forward_reserve=forward_reserve,
    )


def build_workbook(
    decision_df: Optional[pd.DataFrame] = None,
    params: Optional[Dict] = None,
    decision_vars_path: Path = DECISION_VARS_CSV_PATH,
    params_path: Path = PARAMS_PATH,
    output_path: Optional[Path] = None,
) -> Tuple[bytes, Workbook]:
    sections = build_dashboard_sections(decision_df, params, decision_vars_path, params_path)

    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "RawData"
    for r in dataframe_to_rows(sections.raw_df, index=False, header=True):
        ws_data.append(r)

    ws_dash = wb.create_sheet("Storage Units")
    ws_dash["A1"] = "Storage Units"
    ws_dash["A1"].font = Font(size=18, bold=True)

    row_pos = 4
    for entry in sections.store_buy.get("StoreUnits", []):
        loc_df = entry["data"]
        loc = entry["location"]

        ws_dash[f"A{row_pos}"] = f"StoreUnits at {loc}"
        ws_dash[f"A{row_pos}"].font = Font(size=14, bold=True)

        data_start = row_pos + 1
        ws_dash[f"A{data_start}"] = "Storage Type"
        ws_dash[f"B{data_start}"] = "New Storage Units"
        ws_dash[f"C{data_start}"] = "Current Storage Units"

        r = data_start + 1
        for _, row in loc_df.iterrows():
            ws_dash[f"A{r}"] = row["Storage Type"]
            ws_dash[f"B{r}"] = row["New Storage Units"]
            ws_dash[f"C{r}"] = row["Current Storage Units"]
            r += 1

        data_end = r - 1

        chart = BarChart()
        chart.title = f"StoreUnits vs cStoreUnits at {loc}"
        chart.x_axis.title = "Storage Type"
        chart.y_axis.title = "Units"
        chart.dLbls = DataLabelList(showVal=True)

        data = Reference(ws_dash, min_col=2, max_col=3, min_row=data_start, max_row=data_end)
        cats = Reference(ws_dash, min_col=1, min_row=data_start + 1, max_row=data_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws_dash.add_chart(chart, f"E{row_pos}")
        row_pos = data_end + 18

    for entry in sections.store_buy.get("BuyUnits", []):
        loc_df = entry["data"]
        loc = entry["location"]

        ws_dash[f"A{row_pos}"] = f"BuyUnits at {loc}"
        ws_dash[f"A{row_pos}"].font = Font(size=14, bold=True)

        data_start = row_pos + 1
        ws_dash[f"A{data_start}"] = "Storage Type"
        ws_dash[f"B{data_start}"] = "Units"

        r = data_start + 1
        for _, row in loc_df.iterrows():
            ws_dash[f"A{r}"] = row["Storage Type"]
            ws_dash[f"B{r}"] = row["Units"]
            r += 1

        data_end = r - 1

        chart = BarChart()
        chart.title = f"BuyUnits at {loc}"
        chart.x_axis.title = "Storage Type"
        chart.y_axis.title = "Units"
        chart.dLbls = DataLabelList(showVal=True)

        data = Reference(ws_dash, min_col=2, min_row=data_start, max_row=data_end)
        cats = Reference(ws_dash, min_col=1, min_row=data_start + 1, max_row=data_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        ws_dash.add_chart(chart, f"E{row_pos}")
        row_pos = data_end + 18

    relocations = sections.relocations
    if relocations:
        ws_dash[f"A{row_pos}"] = "Relocation Summary"
        ws_dash[f"A{row_pos}"].font = Font(size=16, bold=True)
        row_pos += 2

        for entry in relocations:
            from_loc = entry["from_location"]
            pivot = entry["table"]

            ws_dash[f"A{row_pos}"] = f"Relocations FROM {from_loc}"
            ws_dash[f"A{row_pos}"].font = Font(size=14, bold=True)
            row_pos += 1

            start_table_row = row_pos
            for r in dataframe_to_rows(pivot, index=True, header=True):
                ws_dash.append(r)

            end_table_row = start_table_row + pivot.shape[0] + 1

            chart = BarChart()
            chart.type = "col"
            chart.title = f"Relocations FROM {from_loc}"
            chart.x_axis.title = "Destination Location"
            chart.y_axis.title = "Units Moved"
            chart.legend.position = "r"

            cats = Reference(ws_dash, min_col=1, min_row=start_table_row + 2, max_row=end_table_row)
            for col_idx, storage_type in enumerate(pivot.columns, start=2):
                series = Series(
                    Reference(
                        ws_dash,
                        min_col=col_idx,
                        max_col=col_idx,
                        min_row=start_table_row + 2,
                        max_row=end_table_row,
                    ),
                    title=storage_type,
                )
                chart.series.append(series)

            chart.set_categories(cats)
            chart.dLbls = DataLabelList(showVal=True)
            ws_dash.add_chart(chart, f"E{start_table_row}")

            row_pos = end_table_row + 15

    ws_forward = wb.create_sheet("Forward %")
    ws_forward["A1"] = "SKU Unique and Cubic Percent (Forward)"
    ws_forward["A1"].font = Font(size=16, bold=True)
    row_pos = 3

    for entry in sections.forward:
        sku = entry["sku"]
        pivot_uni = entry["unique"]
        pivot_cub = entry["cubic"]

        ws_forward[f"A{row_pos}"] = f"SKU Group: {sku} (Unique %)"
        ws_forward[f"A{row_pos}"].font = Font(size=14, bold=True)
        row_pos += 1

        pivot_start_uni = row_pos
        for r in dataframe_to_rows(pivot_uni, index=True, header=True):
            ws_forward.append(r)
        pivot_end_uni = pivot_start_uni + len(pivot_uni)

        chart_uni = BarChart()
        chart_uni.type = "col"
        chart_uni.grouping = "stacked"
        chart_uni.overlap = 100
        chart_uni.title = f"{sku}: Unique % Split"
        chart_uni.legend.position = "r"

        cats_uni = Reference(ws_forward, min_col=1, min_row=pivot_start_uni + 2, max_row=pivot_end_uni + 1)
        for col in range(2, pivot_uni.shape[1] + 2):
            series = Series(
                Reference(
                    ws_forward,
                    min_row=pivot_start_uni + 2,
                    max_row=pivot_end_uni + 1,
                    min_col=col,
                    max_col=col,
                ),
                title=ws_forward.cell(row=pivot_start_uni, column=col).value,
            )
            chart_uni.series.append(series)
        chart_uni.set_categories(cats_uni)
        chart_uni.dLbls = DataLabelList(showVal=True)
        ws_forward.add_chart(chart_uni, f"E{pivot_start_uni}")

        ws_forward[f"P{pivot_start_uni - 1}"] = f"SKU Group: {sku} (Cubic %)"
        ws_forward[f"P{pivot_start_uni - 1}"].font = Font(size=14, bold=True)

        pivot_start_cub = pivot_start_uni
        for idx, r in enumerate(dataframe_to_rows(pivot_cub, index=True, header=True)):
            for c_idx, val in enumerate(r, start=16):
                ws_forward.cell(row=pivot_start_cub + idx, column=c_idx, value=val)
        pivot_end_cub = pivot_start_cub + len(pivot_cub)

        chart_cub = BarChart()
        chart_cub.type = "col"
        chart_cub.grouping = "stacked"
        chart_cub.overlap = 100
        chart_cub.title = f"{sku}: Cubic % Split"
        chart_cub.legend.position = "r"

        cats_cub = Reference(ws_forward, min_col=16, min_row=pivot_start_cub + 2, max_row=pivot_end_cub + 1)
        for col in range(17, 17 + pivot_cub.shape[1]):
            series = Series(
                Reference(
                    ws_forward,
                    min_row=pivot_start_cub + 2,
                    max_row=pivot_end_cub + 1,
                    min_col=col,
                    max_col=col,
                ),
                title=ws_forward.cell(row=pivot_start_cub, column=col).value,
            )
            chart_cub.series.append(series)
        chart_cub.set_categories(cats_cub)
        chart_cub.dLbls = DataLabelList(showVal=True)
        ws_forward.add_chart(chart_cub, f"T{pivot_start_cub}")

        row_pos = max(pivot_end_uni, pivot_end_cub) + 15

    ws_fwd_res = wb.create_sheet("Forward+Reserve %")
    ws_fwd_res["A1"] = "SKU Unique and Cubic Percent (Forward + Reserve)"
    ws_fwd_res["A1"].font = Font(size=16, bold=True)
    row_pos = 3

    unique_table_col = 1
    unique_chart_col = "F"
    cubic_table_col = 15
    cubic_chart_col = "T"

    for entry in sections.forward_reserve:
        sku = entry["sku"]
        pivot_uni = entry["unique"]
        pivot_cub = entry["cubic"]

        if not pivot_uni.empty:
            ws_fwd_res[f"A{row_pos}"] = f"{sku} - Unique % (Forward + Reserve)"
            ws_fwd_res[f"A{row_pos}"].font = Font(size=14, bold=True)
            row_pos += 1

            pivot_start_uni = row_pos
            ws_fwd_res.append(["Location"] + list(pivot_uni.columns))
            for idx, row in pivot_uni.iterrows():
                ws_fwd_res.append([idx] + list(row))
            pivot_end_uni = pivot_start_uni + len(pivot_uni)

            chart_uni = BarChart()
            chart_uni.type = "col"
            chart_uni.grouping = "stacked"
            chart_uni.overlap = 100
            chart_uni.title = f"{sku}: Unique % Split (Forward + Reserve)"
            chart_uni.legend.position = "r"

            cats_uni = Reference(
                ws_fwd_res, min_col=unique_table_col, min_row=pivot_start_uni + 1, max_row=pivot_end_uni
            )
            for col in range(unique_table_col + 1, unique_table_col + pivot_uni.shape[1] + 1):
                series = Series(
                    Reference(
                        ws_fwd_res,
                        min_row=pivot_start_uni + 1,
                        max_row=pivot_end_uni,
                        min_col=col,
                        max_col=col,
                    ),
                    title=ws_fwd_res.cell(row=pivot_start_uni, column=col).value,
                )
                chart_uni.series.append(series)
            chart_uni.set_categories(cats_uni)
            chart_uni.dLbls = DataLabelList(showVal=True)
            ws_fwd_res.add_chart(chart_uni, f"{unique_chart_col}{pivot_start_uni}")
        else:
            pivot_end_uni = row_pos

        if not pivot_cub.empty:
            ws_fwd_res[f"{get_column_letter(cubic_table_col)}{pivot_start_uni - 1}"] = (
                f"{sku} - Cubic % (Forward + Reserve)"
            )
            ws_fwd_res[f"{get_column_letter(cubic_table_col)}{pivot_start_uni - 1}"].font = Font(
                size=14, bold=True
            )

            pivot_start_cub = pivot_start_uni
            pivot_cub_reset = pivot_cub.reset_index()
            for i in range(pivot_cub_reset.shape[0] + 1):
                for j in range(pivot_cub_reset.shape[1]):
                    ws_fwd_res.cell(
                        row=pivot_start_cub + i,
                        column=cubic_table_col + j,
                        value=pivot_cub_reset.iloc[i - 1, j] if i > 0 else pivot_cub_reset.columns[j],
                    )
            pivot_end_cub = pivot_start_cub + len(pivot_cub)

            chart_cub = BarChart()
            chart_cub.type = "col"
            chart_cub.grouping = "stacked"
            chart_cub.overlap = 100
            chart_cub.title = f"{sku}: Cubic % Split (Forward + Reserve)"
            chart_cub.legend.position = "r"

            cats_cub = Reference(
                ws_fwd_res, min_col=cubic_table_col, min_row=pivot_start_cub + 1, max_row=pivot_end_cub
            )
            for col in range(cubic_table_col + 1, cubic_table_col + pivot_cub.shape[1] + 1):
                series = Series(
                    Reference(
                        ws_fwd_res,
                        min_row=pivot_start_cub + 1,
                        max_row=pivot_end_cub,
                        min_col=col,
                        max_col=col,
                    ),
                    title=ws_fwd_res.cell(row=pivot_start_cub, column=col).value,
                )
                chart_cub.series.append(series)
            chart_cub.set_categories(cats_cub)
            chart_cub.dLbls = DataLabelList(showVal=True)
            ws_fwd_res.add_chart(chart_cub, f"{cubic_chart_col}{pivot_start_cub}")
        else:
            pivot_end_cub = pivot_end_uni

        row_pos = max(pivot_end_uni, pivot_end_cub) + 15

    if output_path:
        wb.save(output_path)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue(), wb


if __name__ == "__main__":
    build_workbook(output_path=BASE_DIR / "Dashboard.xlsx")
    print("Dashboard exported to Dashboard.xlsx")
