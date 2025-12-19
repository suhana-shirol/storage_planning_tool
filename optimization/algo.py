from pyomo.environ import *
import pyomo.environ as pyo
import pandas as pd
import numpy as np
import math
import itertools
from typing import List, Dict
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference, Series
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter


pd.set_option('future.no_silent_downcasting', True)

from .retrieve import *

BASE_DIR = Path(__file__).resolve().parent
PIECEWISE_PARAMS_PATH = BASE_DIR / "piecewiseparams.json"
DECISION_VARS_XLSX_PATH = BASE_DIR / "non_zero_decision_variables.xlsx"
DECISION_VARS_CSV_PATH = BASE_DIR / "non_zero_decision_variables.csv"


solver = pyo.SolverFactory('glpk')
# solver = SolverFactory('cbc', executable=r"C:\cbc\bin\cbc.exe")
m = pyo.ConcreteModel()

# indices --------------------------------------------------------------------

sets_dict = build_sets_from_db()
priority = sets_dict['priority']
movement = sets_dict['movement']
sizes = sets_dict['sizes']
PRIORITYSET = sets_dict['PRIORITYSET']
SKUS = sets_dict['SKUS']
USERS = sets_dict['USERS']
LOCS = sets_dict['LOCS']
TYPES = sets_dict['TYPES']
PRIORITIES = sets_dict['PRIORITIES']
BREAKS = [0,1,2]
not_reserve_users = sets_dict['not_reserve_users']
SECONDARY = sets_dict['secondary']
indices = {}
indices['isku'] = SKUS
indices['iuser'] = USERS
indices['iloc'] = LOCS
indices['itype'] = TYPES
indices['ilocpriority'] = PRIORITIES
indices['ibreak'] = BREAKS
indices['secondary'] = SECONDARY


# decison variables ------------------------------------------------------------------------

m.isku = pyo.Set(initialize=indices['isku'])
m.iuser = pyo.Set(initialize=indices['iuser'])
m.iloc = pyo.Set(initialize=indices['iloc'])
m.itype = pyo.Set(initialize=indices['itype'])
m.ilocpriority = pyo.Set(initialize=indices['ilocpriority'])
m.ibreak = pyo.Set(initialize=indices['ibreak'])
m.secondary = pyo.Set(initialize=indices['secondary'])


m.StoreUnits = pyo.Var(m.itype, m.iloc, within = pyo.NonNegativeIntegers)
m.BuyUnits = pyo.Var(m.itype, m.iloc, within = pyo.NonNegativeIntegers)
m.RelocUnits = pyo.Var(m.itype, m.iloc,m.iloc, within = pyo.NonNegativeIntegers)

m.SKUStoreUnits = pyo.Var(m.isku, m.itype, m.iloc, within = pyo.NonNegativeReals)
m.SKUcubicPct = pyo.Var(m.isku, m.itype, m.iloc, within = pyo.NonNegativeReals)
m.SKUuniquePct = pyo.Var(m.isku, m.itype, m.iloc, within = pyo.NonNegativeReals)
m.ResSKUcubicPct = pyo.Var(m.isku, m.itype, m.iloc, within = pyo.NonNegativeReals)
m.ResSKUuniquePct = pyo.Var(m.isku, m.itype, m.iloc, within = pyo.NonNegativeReals)
m.TotalCubicPct = pyo.Var(m.isku,m.itype, m.iloc, within = pyo.NonNegativeReals)

m.AllocFracUnique = pyo.Var(m.isku, m.itype, m.iloc, m.iuser, within = pyo.NonNegativeReals)

m.LineDownOrders = pyo.Var(m.isku, m.iuser, m.ilocpriority, within = pyo.NonNegativeReals)
m.IncLDOrders = pyo.Var(m.isku, m.iuser, m.ilocpriority, within = pyo.NonNegativeReals)
m.x = pyo.Var(m.isku, m.iuser, m.ilocpriority, within = pyo.NonNegativeReals)
m.Lambda = pyo.Var(m.isku, m.iuser, m.ilocpriority, m.ibreak, within = pyo.NonNegativeReals)
m.upto = pyo.Var(m.isku, m.iuser, m.ilocpriority, m.ibreak, within = pyo.Binary)


m.reserveorders = pyo.Var(m.isku, m.iloc, within = pyo.NonNegativeReals)
m.z = pyo.Var(m.isku, m.iloc, within = pyo.NonNegativeReals)
m.theta = pyo.Var(m.isku, m.iloc, m.ibreak, within = pyo.NonNegativeReals)
m.resupto = pyo.Var(m.isku, m.iloc, m.ibreak, within = pyo.Binary)




#parameters ------------------------------------------------------------------------------------

i_loc = get_i_loc()
i_locpriority = get_i_locpriority()
i_sku = get_i_sku()
i_sku_type = get_i_sku_type()
i_type = get_i_type()
i_type_loc = get_i_type_loc()
i_user_loc = get_i_loc_user()
i_sku_user = get_i_sku_user()


all_params = {'iloc': i_loc,
              'ilocpriority': i_locpriority,
              'isku': i_sku,
              'iskutype': i_sku_type,
              'itype': i_type,
              'itypeloc': i_type_loc,
              'i_user_loc': i_user_loc,
              'i_sku_user': i_sku_user
              }

Params = {k: v for outer in all_params.values() for k, v in outer.items()}

system_params = get_system()
Params['Linedowncost'] = system_params['Linedowncost'].iloc[0]
Params['Budget'] = system_params['Budget'].iloc[0]
Params['SpaceFactor'] = system_params['SpaceFactor'].iloc[0]
Params['reservecost'] = system_params['reservecost'].iloc[0]


if not PIECEWISE_PARAMS_PATH.exists():
    raise FileNotFoundError(f"Piecewise params not found at {PIECEWISE_PARAMS_PATH}")

with PIECEWISE_PARAMS_PATH.open("r") as f:
    piecewiseparams = json.load(f)

def strings_to_tuples(d):
    return {eval(k): v for k, v in d.items()}

piecewiseparams = {
    "breakpoints": strings_to_tuples(piecewiseparams["breakpoints"]),
    "pieceval": strings_to_tuples(piecewiseparams["pieceval"]),
    "resbreakpoints": strings_to_tuples(piecewiseparams["resbreakpoints"]),
    "respieceval": strings_to_tuples(piecewiseparams["respieceval"]),
}
Params.update(piecewiseparams)


def _json_safe(o):
    if isinstance(o, dict):
        return {str(k): _json_safe(v) for k, v in o.items()}
    if isinstance(o, (list, tuple, set)):
        return [_json_safe(v) for v in o]
    if hasattr(o, "item"):  # numpy scalar
        return o.item()
    return o

DEBUG_PARAMS_PATH = Path(__file__).resolve().parent / "params.json"
with DEBUG_PARAMS_PATH.open("w") as f:
    json.dump(_json_safe(Params), f, indent=2)
print(f"Wrote params to {DEBUG_PARAMS_PATH}")

#objective  ---------------------------------------------------------------------------------------

m.Obj = pyo.Objective(
    expr=(
        sum(m.BuyUnits[typ, loc] * Params['BuyInvest'][typ]
            for typ in m.itype for loc in m.iloc)

        + sum(m.RelocUnits[typ, from_loc, to_loc] * Params['RelocInvest'][typ] / 2
              for typ in m.itype for from_loc in m.iloc for to_loc in m.iloc if from_loc != to_loc)

        + sum(Params['Linedowncost'] * Params['traveltime'][locpriority] *
              m.IncLDOrders[sku, user, locpriority]
              for sku in m.isku for user in m.iuser for locpriority in m.ilocpriority)

        + sum((Params['reservecost'] / (Params["LineDownOrders"][sku] + 1)) * m.reserveorders[sku, loc]
              for sku in m.isku for loc in m.iloc)

        + sum(Params['penalty'][sku, typ] * m.SKUStoreUnits[sku, typ, loc] for sku in m.isku for typ in m.itype for loc in m.iloc)
    ),
    sense=pyo.minimize
)


#constraints ---------------------------------------------------------------------------------------

#1.1 - Restrict locations to users
m.UserLocCompat = pyo.ConstraintList()
for sku in m.isku:
    for loc in m.iloc:
        for typ in m.itype:
            for user in m.iuser:
                m.UserLocCompat.add(
                    m.AllocFracUnique[sku, typ, loc, user] <= Params['UserLocationCompat'][loc, user])

#1.2 - Restrict allocation of skus to storage types
m.TypeLocCompat = pyo.ConstraintList()
for sku in m.isku:
    for loc in m.iloc:
        for typ in m.itype:
            for user in m.iuser:
                m.TypeLocCompat.add(
                    m.AllocFracUnique[sku, typ, loc, user] <= Params['compatible'][sku, typ])

#1.3 - Restrict storage types to sku groups
m.skutypeCompat = pyo.ConstraintList()
for sku in m.isku:
    for loc in m.iloc:
        for typ in m.itype:
            m.TypeLocCompat.add(
                m.SKUStoreUnits[sku, typ, loc] <=1000* Params['compatible'][sku, typ])





#2.1 - Storage unit flows
m.StorageUnitFlows = pyo.ConstraintList()
for typ in m.itype:
    for loc in m.iloc:
        inbound = sum(m.RelocUnits[typ, jloc, loc] for jloc in m.iloc if jloc != loc)
        outbound = sum(m.RelocUnits[typ, loc, jloc] for jloc in m.iloc if jloc != loc)

        m.StorageUnitFlows.add(
            m.StoreUnits[typ, loc] == Params['cStoreUnits'][typ, loc]
                                   + m.BuyUnits[typ, loc]
                                   + inbound - outbound)

#2.2 - Storage unit movement restriction
m.NoSelfRelocation = pyo.ConstraintList()
for loc in m.iloc:
    for typ in m.itype:
        m.NoSelfRelocation.add(m.RelocUnits[typ, loc, loc] == 0)

#2.3 - Budget constraint
m.BudgetConstraint = pyo.ConstraintList()
m.BudgetConstraint.add(
    sum(m.BuyUnits[typ, loc] * Params['BuyExpense'][typ]for loc in m.iloc for typ in m.itype)
    + sum(m.RelocUnits[typ, from_loc, to_loc] * Params['RelocExpense'][typ] / 2
           for typ in m.itype
           for from_loc in m.iloc
           for to_loc in m.iloc
           if from_loc != to_loc) <= Params['Budget'])

#2.4 - Location space capacity
m.LocationSpaceCap = pyo.ConstraintList()
for loc in m.iloc:
    m.LocationSpaceCap.add(
        sum(m.StoreUnits[typ, loc] * Params['SpaceReq'][typ]
            for typ in m.itype)
        <= Params['FloorSpace'][loc] * Params['SpaceFactor'])

#2.5 - Relationship of store units
m.StoreToSKUUnits = pyo.ConstraintList()
for typ in m.itype:
    for loc in m.iloc:
        m.StoreToSKUUnits.add(
            m.StoreUnits[typ, loc] ==sum(m.SKUStoreUnits[sku, typ, loc] for sku in m.isku))

#3.1 - Storage unit cubic capacity per sku group
m.StorageUnitFt = pyo.ConstraintList()
for sku in m.isku:
    for typ in m.itype:
        for loc in m.iloc:
            m.StorageUnitFt.add(
                m.SKUStoreUnits[sku, typ, loc] * Params['cubiccapperunit'][sku, typ]
                * Params['VertSpace'][typ, loc]
                >= Params['SKUcubicft'][sku] *
                   (m.SKUcubicPct[sku, typ, loc] + m.ResSKUcubicPct[sku, typ, loc]))

#3.2 - volume and unique sku connection
m.VolumeUniqueConnect = pyo.ConstraintList()
for sku in m.isku:
    for typ in m.itype:
        for loc in m.iloc:
            m.VolumeUniqueConnect.add(
                Params['SKUcubicft'][sku] * m.SKUcubicPct[sku, typ, loc]
                >= m.SKUuniquePct[sku, typ, loc]
                   * Params['NumSKU'][sku]
                   * Params['minVol'][sku])

m.ResVolumeUniqueConnect = pyo.ConstraintList()
for sku in m.isku:
    for typ in m.itype:
        for loc in m.iloc:
            m.ResVolumeUniqueConnect.add(
                Params['SKUcubicft'][sku] * m.ResSKUcubicPct[sku, typ, loc]
                >= m.ResSKUuniquePct[sku, typ, loc]
                   * Params['NumSKU'][sku]
                   * Params['minVol'][sku])

#3.3 - unique sku per storage unit upper bound
m.StorageUnitCapacityUnique = pyo.ConstraintList()
for sku in m.isku:
    for typ in m.itype:
        for loc in m.iloc:
            m.StorageUnitCapacityUnique.add(
                m.SKUStoreUnits[sku, typ, loc] * Params['maxskusperunit'][sku, typ]
                >= Params['NumSKU'][sku] * m.SKUuniquePct[sku, typ, loc])

m.ResStorageUnitCapacityUnique = pyo.ConstraintList()
for sku in m.isku:
    for typ in m.itype:
        for loc in m.iloc:
            m.ResStorageUnitCapacityUnique.add(
                m.SKUStoreUnits[sku, typ, loc] * Params['maxskusperunit'][sku, typ]
                >= Params['NumSKU'][sku] * m.ResSKUuniquePct[sku, typ, loc])


#3.4 - cubic ft allocated upperbound
m.CubicPctBoundUnique = pyo.ConstraintList()
for sku in m.isku:
    for typ in m.itype:
        for loc in m.iloc:
            m.CubicPctBoundUnique.add(
                m.SKUcubicPct[sku, typ, loc] <= 1.2 * m.SKUuniquePct[sku, typ, loc])

m.ResCubicPctBoundUnique = pyo.ConstraintList()
for sku in m.isku:
    for typ in m.itype:
        for loc in m.iloc:
            m.ResCubicPctBoundUnique.add(
                m.ResSKUcubicPct[sku, typ, loc] <= 1.2 * m.ResSKUuniquePct[sku, typ, loc])


# # 3.5 new
# m.CubicPctBoundUnique = pyo.ConstraintList()
# for sku in m.isku:
#     for typ in m.itype:
#         for loc in m.iloc:
#             m.CubicPctBoundUnique.add(
#                 1000 * m.SKUcubicPct[sku, typ, loc] >=  m.SKUuniquePct[sku, typ, loc])

# m.ResCubicPctBoundUnique = pyo.ConstraintList()
# for sku in m.isku:
#     for typ in m.itype:
#         for loc in m.iloc:
#             m.ResCubicPctBoundUnique.add(
#                 1000 * m.ResSKUcubicPct[sku, typ, loc] >=  m.ResSKUuniquePct[sku, typ, loc])





#4.1 - sku cubic footage coverage
m.CubicSKUAlloc = pyo.ConstraintList()
for sku in m.isku:
    m.CubicSKUAlloc.add(
        sum(m.SKUcubicPct[sku, typ, loc] + m.ResSKUcubicPct[sku, typ, loc]
            for typ in m.itype for loc in m.iloc) == 1)


#4.2 - sku unique percent coverage
m.UniqueSKUAlloc = pyo.ConstraintList()

for sku in m.isku:
    m.UniqueSKUAlloc.add(
        sum(m.SKUuniquePct[sku, typ, loc]
            for typ in m.itype for loc in m.iloc) >= 1)


#4.3 - Aligning unique sku allocation to users
m.UserAlloc = pyo.ConstraintList()
for sku in m.isku:
    for typ in m.itype:
        for loc in m.iloc:
            m.UserAlloc.add(
                sum(m.AllocFracUnique[sku, typ, loc, user] * Params['NumSKUUser'][sku, user]
                    for user in m.iuser)
                >= m.SKUuniquePct[sku, typ, loc] * Params['NumSKU'][sku])

#4.4 - Cannot allocate more skus than what is there
m.UserAllocLimit = pyo.ConstraintList()
for sku in m.isku:
    for typ in m.itype:
        for loc in m.iloc:
            for user in m.iuser:
                m.UserAllocLimit.add(
                    m.AllocFracUnique[sku, typ, loc, user] * Params['NumSKUUser'][sku, user]
                    <= m.SKUuniquePct[sku, typ, loc] * Params['NumSKU'][sku]
                )

# 4.5 - Percent of user's skus must all be allocated
m.UserAllocation = pyo.ConstraintList()
for sku in m.isku:
    for user in m.iuser:
        m.UserAllocation.add(
            sum(m.AllocFracUnique[sku, typ, loc, user]
                for typ in m.itype for loc in m.iloc) == 1)


#4.6 - Number of skus for a user must be less than or equal to total skus
m.UserCapacityLimit = pyo.ConstraintList()
for sku in m.isku:
    for typ in m.itype:
        for loc in m.iloc:
            m.UserCapacityLimit.add(
                sum(m.AllocFracUnique[sku, typ, loc, user] * Params['NumSKUUser'][sku, user]
                    for user in m.iuser)
                <= sum(Params['NumSKU'][sku] for user in m.iuser)* m.SKUuniquePct[sku, typ, loc])


#5.1
m.ReserveLocationLimit = pyo.ConstraintList()
for sku in m.isku:
    for typ in m.itype:
        for loc in m.iloc:
            m.ReserveLocationLimit.add(
                m.ResSKUuniquePct[sku, typ, loc] <= Params['ReserveAllowed'][loc]
            )

m.ReserveCubicLocationLimit = pyo.ConstraintList()
for sku in m.isku:
    for typ in m.itype:
        for loc in m.iloc:
            m.ReserveCubicLocationLimit.add(
                m.ResSKUcubicPct[sku, typ, loc] <= Params['ReserveAllowed'][loc])


#5.2
# m.TotalCubicDefinition = pyo.ConstraintList()
# for sku in m.isku:
#     for typ in m.itype:
#         for loc in m.iloc:
#             m.TotalCubicDefinition.add(
#                 m.SKUcubicPct[sku, typ, loc] +
#                 m.ResSKUcubicPct[sku, typ, loc] ==
#                 m.TotalCubicPct[sku, typ, loc])



# 6.1 - Line down orders covered for each location priority category

def lambda_sum_rule(m, sku, user, prio):
    return sum(m.Lambda[sku, user, prio, brk] for brk in m.ibreak) == 1
m.LambdaSum = pyo.Constraint(m.isku, m.iuser, m.ilocpriority, rule=lambda_sum_rule)

def upto_rule(m, sku, user, prio):
    return m.upto[sku, user, prio, 1] + m.upto[sku, user, prio, 2] == 1
m.UptoFirst = pyo.Constraint(m.isku, m.iuser, m.ilocpriority, rule=upto_rule)


def lambda0_upto1_rule(m, sku, user, prio):
    return m.Lambda[sku,user,prio,0] <= m.upto[sku,user,prio,1]
m.Lambda0Upto1 = pyo.Constraint(m.isku, m.iuser, m.ilocpriority, rule=lambda0_upto1_rule)


def lambda1_upto_rule(m, sku, user, prio):
    return m.Lambda[sku,user,prio,1] <= m.upto[sku,user,prio,1] + m.upto[sku,user,prio,2]
m.Lambda1Upto = pyo.Constraint(m.isku, m.iuser, m.ilocpriority, rule=lambda1_upto_rule)

def lambda2_upto2_rule(m, sku, user, prio):
    return m.Lambda[sku,user,prio,2] <= m.upto[sku,user,prio,2]
m.Lambda2Upto2 = pyo.Constraint(m.isku, m.iuser, m.ilocpriority, rule=lambda2_upto2_rule)

def x_relationship_rule(m, sku, user, prio):
    return m.x[sku,user,prio] == sum(
        m.Lambda[sku,user,prio,brk] * Params['breakpoints'][sku,user,brk]
        for brk in m.ibreak
    )
m.XRelationship = pyo.Constraint(m.isku, m.iuser, m.ilocpriority, rule=x_relationship_rule)

def x_total_rule(m, sku, user, prio):
    return m.x[sku,user,prio] == sum(
        m.AllocFracUnique[sku, typ, loc, user] * Params['NumSKUUser'][(sku,user)]
        for loc in PRIORITYSET[prio] for typ in m.itype
    )
m.XTotal = pyo.Constraint(m.isku, m.iuser, m.ilocpriority, rule=x_total_rule)

def line_down_rule(m, sku, user, prio):
    return m.LineDownOrders[sku,user,prio] == sum(
        m.Lambda[sku,user,prio,brk] * Params['pieceval'][sku,user,brk]
        for brk in m.ibreak
    )
m.LDOrdersDef = pyo.Constraint(m.isku, m.iuser, m.ilocpriority, rule=line_down_rule)

def inc_ld_rule(m, sku, user, prio):
    if prio == 1:
        return m.IncLDOrders[sku,user,prio] == m.LineDownOrders[sku,user,prio]
    return m.IncLDOrders[sku,user,prio] == (
        m.LineDownOrders[sku,user,prio] - m.LineDownOrders[sku,user,prio - 1]
    )
m.IncLDConstraint = pyo.Constraint(m.isku, m.iuser, m.ilocpriority, rule=inc_ld_rule)



#6.2 Piecewise linear function for reserve orders

def theta_sum_rule(m, sku, loc):
    return sum(m.theta[sku, loc, brk] for brk in m.ibreak) == 1
m.ThetaSum = pyo.Constraint(m.isku, m.secondary, rule=theta_sum_rule)

def resupto_rule(m, sku, loc):
    return m.resupto[sku, loc, 1] + m.resupto[sku, loc, 2] == 1
m.ResUptoFirst = pyo.Constraint(m.isku, m.secondary, rule=resupto_rule)

def theta0_upto1_rule(m, sku, loc):
    return m.theta[sku, loc, 0] <= m.resupto[sku, loc, 1]
m.Theta0Upto1 = pyo.Constraint(m.isku, m.secondary, rule=theta0_upto1_rule)

def theta1_upto_rule(m, sku, loc):
    return m.theta[sku, loc, 1] <= (
        m.resupto[sku, loc, 1] + m.resupto[sku, loc, 2]
    )
m.Theta1Upto = pyo.Constraint(m.isku, m.secondary, rule=theta1_upto_rule)

def theta2_upto2_rule(m, sku, loc):
    return m.theta[sku, loc, 2] <= m.resupto[sku, loc, 2]
m.Theta2Upto2 = pyo.Constraint(m.isku, m.secondary, rule=theta2_upto2_rule)

def z_relationship_rule(m, sku, loc):
    return m.z[sku, loc] == sum(
        m.theta[sku, loc, brk] * Params['resbreakpoints'][sku, brk]
        for brk in m.ibreak
    )
m.ZRelationship = pyo.Constraint(m.isku, m.secondary, rule=z_relationship_rule)

def z_total_rule(m, sku, loc):
    return m.z[sku, loc] == sum(
        m.ResSKUuniquePct[sku, typ, loc] * Params['NumSKU'][sku]
        for typ in m.itype
    )
m.ZTotal = pyo.Constraint(m.isku, m.secondary, rule=z_total_rule)

def reserve_orders_rule(m, sku, loc):
    return m.reserveorders[sku, loc] >= sum(
        m.theta[sku, loc, brk] * Params['respieceval'][sku, brk]
        for brk in m.ibreak
    )
m.ResOrdersDef = pyo.Constraint(m.isku, m.secondary, rule=reserve_orders_rule)


#solve ------------------------------------------------------------------------------
solver.options['tmlim'] = 150
# solver.options['mipgap'] = 0.022
solution = solver.solve(m, tee=True)

print("Status:", solution.solver.status)

data = []
for var in m.component_data_objects(pyo.Var):
    if var.value is not None and abs(var.value) > 1e-6:
        var_name = var.name

        if "[" in var_name:
            base_name = var_name.split("[")[0]
            indices = var_name[var_name.find("[")+1 : var_name.find("]")].split(",")
        else:
            base_name = var_name
            indices = []

        entry = {"Variable": base_name, "Value": var.value}
        for i, idx in enumerate(indices):
            entry[f"Index_{i+1}"] = idx.strip()

        data.append(entry)
df = pd.DataFrame(data)

# Ensure Value column exists even if no data (keeps reindex safe)
if "Value" not in df.columns:
    df["Value"] = np.nan

# Reorder columns: all index/meta cols first, Value last
cols = [col for col in df.columns if col != "Value"] + ["Value"]
df = df.loc[:, cols]

# Append objective value row for dashboard consumption
# total_buy = sum(m.BuyUnits[typ, loc] * Params['BuyInvest'][typ]
#             for typ in m.itype for loc in m.iloc)
# if obj_val is not None:
#     obj_row = {col: "" for col in df.columns}
#     obj_row["Variable"] = "objective value"
#     obj_row["Value"] = obj_val
#     df = pd.concat([df, pd.DataFrame([obj_row])], ignore_index=True)


df.to_excel(DECISION_VARS_XLSX_PATH, index=False)
print(f"Non-zero decision variables exported to {DECISION_VARS_XLSX_PATH}")


df.to_excel(DECISION_VARS_CSV_PATH, index=False)
print(f"Non-zero decision variables exported to {DECISION_VARS_CSV_PATH}")


# =========================== DASHBOARD CREATION ===============================
chart_variables = ["StoreUnits", "BuyUnits"]

wb = Workbook()

ws_data = wb.active
ws_data.title = "RawData"
for r in dataframe_to_rows(df, index=False, header=True):
    ws_data.append(r)

ws_dash = wb.create_sheet("Storage Units")
ws_dash["A1"] = "Storage Units"
ws_dash["A1"].font = Font(size=18, bold=True)

row_pos = 4


cstore_df = pd.DataFrame(
    [(k[0], k[1], v) for k, v in Params["cStoreUnits"].items()],
    columns=["Index_1", "Index_2", "cStoreUnits"]
).dropna()

# =================== STOREUNITS & BUYUNITS ===================
for var in chart_variables:
    var_df = df[df["Variable"] == var]
    if var_df.empty:
        continue

    for loc in var_df["Index_2"].unique():
        loc_df = var_df[var_df["Index_2"] == loc]

        ws_dash[f"A{row_pos}"] = f"{var} at {loc}"
        ws_dash[f"A{row_pos}"].font = Font(size=14, bold=True)

        data_start = row_pos + 1

        if var == "StoreUnits":
            merged_df = loc_df.merge(
                cstore_df,
                on=["Index_1", "Index_2"],
                how="left"
            ).fillna(0).infer_objects(copy=False)

            ws_dash[f"A{data_start}"] = "Storage Type"
            ws_dash[f"B{data_start}"] = "New Storage Units"
            ws_dash[f"C{data_start}"] = "Current Storage Units"

            r = data_start + 1
            for _, row in merged_df.iterrows():
                ws_dash[f"A{r}"] = row["Index_1"]
                ws_dash[f"B{r}"] = row["Value"]
                ws_dash[f"C{r}"] = row["cStoreUnits"]
                r += 1

            data_end = r - 1

            chart = BarChart()
            chart.title = f"{var} vs cStoreUnits at {loc}"
            chart.x_axis.title = "Storage Type"
            chart.y_axis.title = "Units"
            chart.dLbls = DataLabelList(showVal=True)

            data = Reference(ws_dash, min_col=2, max_col=3, min_row=data_start, max_row=data_end)
            cats = Reference(ws_dash, min_col=1, min_row=data_start + 1, max_row=data_end)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

        else:
            ws_dash[f"A{data_start}"] = "Storage Type"
            ws_dash[f"B{data_start}"] = "Units"

            r = data_start + 1
            for _, row in loc_df.iterrows():
                ws_dash[f"A{r}"] = row["Index_1"]
                ws_dash[f"B{r}"] = row["Value"]
                r += 1

            data_end = r - 1

            chart = BarChart()
            chart.title = f"{var} at {loc}"
            chart.x_axis.title = "Storage Type"
            chart.y_axis.title = "Units"
            chart.dLbls = DataLabelList(showVal=True)

            data = Reference(ws_dash, min_col=2, min_row=data_start, max_row=data_end)
            cats = Reference(ws_dash, min_col=1, min_row=data_start + 1, max_row=data_end)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)

        ws_dash.add_chart(chart, f"E{row_pos}")
        row_pos = data_end + 18


# =================== RELOCATIONS SECTION =========================
reloc_df = df[df["Variable"] == "RelocUnits"]
if not reloc_df.empty:

    ws_dash[f"A{row_pos}"] = "Relocation Summary"
    ws_dash[f"A{row_pos}"].font = Font(size=16, bold=True)
    row_pos += 2

    reloc_df = reloc_df.copy()

    reloc_df["Index_3"] = reloc_df["Index_3"].fillna("Unknown").infer_objects(copy=False)

    from_locs = reloc_df["Index_2"].unique()

    for from_loc in from_locs:
        ws_dash[f"A{row_pos}"] = f"Relocations FROM {from_loc}"
        ws_dash[f"A{row_pos}"].font = Font(size=14, bold=True)
        row_pos += 1

        df_from = reloc_df[reloc_df["Index_2"] == from_loc]

        pivot = df_from.pivot_table(
            index="Index_3",
            columns="Index_1",
            values="Value",
            aggfunc="sum",
            fill_value=0
        )
        pivot.index.name = ""

        start_table_row = row_pos

        for r in dataframe_to_rows(pivot, index=True, header=True):
            ws_dash.append(r)

        end_table_row = start_table_row + pivot.shape[0] + 1

        chart = BarChart()
        chart.type = "col"
        chart.title = f"Relocations FROM {from_loc}"
        chart.x_axis.title = "Destination Location"
        chart.y_axis.title = "Units Moved"
        chart.legend.position = "r"

        cats = Reference(ws_dash,
                         min_col=1,
                         min_row=start_table_row + 2,
                         max_row=end_table_row)

        for col_idx, storage_type in enumerate(pivot.columns, start=2):
            series = Series(
                Reference(ws_dash,
                          min_col=col_idx, max_col=col_idx,
                          min_row=start_table_row + 2, max_row=end_table_row),
                title=storage_type
            )
            chart.series.append(series)

        chart.set_categories(cats)
        chart.dLbls = DataLabelList(showVal=True)

        # Place chart
        ws_dash.add_chart(chart, f"E{start_table_row}")

        row_pos = end_table_row + 15


# --------------------------------------------FORWARD UNIQUE + CUBIC---------------------------------------------------------------

ws_sku = wb.create_sheet("Forward %")
ws_sku["A1"] = "SKU Unique and Cubic Percent (Forward)"
ws_sku["A1"].font = Font(size=16, bold=True)

row_pos = 3

# Load data
uni_df = df[df["Variable"] == "SKUuniquePct"].copy()
cub_df = df[df["Variable"] == "SKUcubicPct"].copy()

uni_df["Value"] = uni_df["Value"].replace(0, None).apply(lambda x: 0.01 if 0 < x < 0.01 else round(x, 2))
cub_df["Value"] = cub_df["Value"].replace(0, None).apply(lambda x: 0.01 if 0 < x < 0.01 else round(x, 2))

for sku in uni_df["Index_1"].unique():

    # ======================= LEFT SIDE: UNIQUE ==========================
    group_uni = uni_df[uni_df["Index_1"] == sku].dropna(subset=["Value"])
    pivot_uni = group_uni.pivot_table(index="Index_3", columns="Index_2",
                                      values="Value", aggfunc="sum", fill_value=0)
    pivot_uni.index.name = ""

    ws_sku[f"A{row_pos}"] = f"SKU Group: {sku} (Unique %)"
    ws_sku[f"A{row_pos}"].font = Font(size=14, bold=True)
    row_pos += 1

    pivot_start_uni = row_pos

    for r in dataframe_to_rows(pivot_uni, index=True, header=True):
        ws_sku.append(r)

    pivot_end_uni = pivot_start_uni + len(pivot_uni)

    chart_uni = BarChart()
    chart_uni.type = "col"
    chart_uni.grouping = "stacked"
    chart_uni.overlap = 100
    chart_uni.title = f"{sku}: Unique % Split"
    chart_uni.legend.position = "r"

    cats_uni = Reference(ws_sku, min_col=1, min_row=pivot_start_uni + 2, max_row=pivot_end_uni + 1)

    for col in range(2, pivot_uni.shape[1] + 2):
        series = Series(
            Reference(ws_sku, min_row=pivot_start_uni + 2, max_row=pivot_end_uni + 1,
                     min_col=col, max_col=col),
            title=ws_sku.cell(row=pivot_start_uni, column=col).value
        )
        chart_uni.series.append(series)

    chart_uni.set_categories(cats_uni)
    chart_uni.dLbls = DataLabelList(showVal=True)

    ws_sku.add_chart(chart_uni, f"E{pivot_start_uni}")

    # ======================= RIGHT SIDE: CUBIC ==========================
    group_cub = cub_df[cub_df["Index_1"] == sku].dropna(subset=["Value"])
    pivot_cub = group_cub.pivot_table(index="Index_3", columns="Index_2",
                                      values="Value", aggfunc="sum", fill_value=0)
    pivot_cub.index.name = ""

    ws_sku[f"P{pivot_start_uni - 1}"] = f"SKU Group: {sku} (Cubic %)"
    ws_sku[f"P{pivot_start_uni - 1}"].font = Font(size=14, bold=True)

    pivot_start_cub = pivot_start_uni

    for idx, r in enumerate(dataframe_to_rows(pivot_cub, index=True, header=True)):
        for c_idx, val in enumerate(r, start=16):
            ws_sku.cell(row=pivot_start_cub + idx, column=c_idx, value=val)

    pivot_end_cub = pivot_start_cub + len(pivot_cub)

    chart_cub = BarChart()
    chart_cub.type = "col"
    chart_cub.grouping = "stacked"
    chart_cub.overlap = 100
    chart_cub.title = f"{sku}: Cubic % Split"
    chart_cub.legend.position = "r"

    cats_cub = Reference(ws_sku, min_col=16, min_row=pivot_start_cub + 2, max_row=pivot_end_cub + 1)

    for col in range(17, 17 + pivot_cub.shape[1]):
        series = Series(
            Reference(ws_sku, min_row=pivot_start_cub + 2, max_row=pivot_end_cub + 1,
                     min_col=col, max_col=col),
            title=ws_sku.cell(row=pivot_start_cub, column=col).value
        )
        chart_cub.series.append(series)

    chart_cub.set_categories(cats_cub)
    chart_cub.dLbls = DataLabelList(showVal=True)

    ws_sku.add_chart(chart_cub, f"T{pivot_start_cub}")

    row_pos = max(pivot_end_uni, pivot_end_cub) + 15



#========================================== FORWARD + RESERVE =============================
ws_sku = wb.create_sheet("Forward+Reserve %")
ws_sku["A1"] = "SKU Unique and Cubic Percent (Forward + Reserve)"
ws_sku["A1"].font = Font(size=16, bold=True)

row_pos = 3

uni_reg = df[df["Variable"] == "SKUuniquePct"].copy()
uni_res = df[df["Variable"] == "ResSKUuniquePct"].copy()

cub_reg = df[df["Variable"] == "SKUcubicPct"].copy()
cub_res = df[df["Variable"] == "ResSKUcubicPct"].copy()


for d in [uni_res, cub_res]:
    if not d.empty:
        d["Index_2"] = d["Index_2"] + "_Reserve"


uni_df = pd.concat([uni_reg, uni_res], ignore_index=True)
cub_df = pd.concat([cub_reg, cub_res], ignore_index=True)

uni_df["Value"] = uni_df["Value"].replace(0, None).apply(lambda x: 0.01 if 0 < x < 0.01 else round(x, 2))
cub_df["Value"] = cub_df["Value"].replace(0, None).apply(lambda x: 0.01 if 0 < x < 0.01 else round(x, 2))


unique_table_col = 1
unique_chart_col = "F"
cubic_table_col = 15
cubic_chart_col = "T"


for sku in uni_df["Index_1"].unique():
    # ======================= LEFT SIDE: UNIQUE ============================
    group_uni = uni_df[uni_df["Index_1"] == sku].dropna(subset=["Value"])
    if not group_uni.empty:
        pivot_uni = group_uni.pivot_table(index="Index_3", columns="Index_2",
                                          values="Value", aggfunc="sum", fill_value=0)
        pivot_uni.index.name = ""

        ws_sku[f"A{row_pos}"] = f"{sku} - Unique % (Forward + Reserve)"
        ws_sku[f"A{row_pos}"].font = Font(size=14, bold=True)
        row_pos += 1

        pivot_start_uni = row_pos

        ws_sku.append(["Location"] + list(pivot_uni.columns))
        for idx, row in pivot_uni.iterrows():
            ws_sku.append([idx] + list(row))
        pivot_end_uni = pivot_start_uni + len(pivot_uni)

        # Unique Chart
        chart_uni = BarChart()
        chart_uni.type = "col"
        chart_uni.grouping = "stacked"
        chart_uni.overlap = 100
        chart_uni.title = f"{sku}: Unique % Split (Forward + Reserve)"
        chart_uni.legend.position = "r"

        cats_uni = Reference(ws_sku, min_col=unique_table_col, min_row=pivot_start_uni + 1,
                             max_row=pivot_end_uni )

        for col in range(unique_table_col + 1, unique_table_col + pivot_uni.shape[1] + 1):
            series = Series(
                Reference(ws_sku, min_row=pivot_start_uni + 1, max_row=pivot_end_uni ,
                         min_col=col, max_col=col),
                title=ws_sku.cell(row=pivot_start_uni, column=col).value
            )
            chart_uni.series.append(series)

        chart_uni.set_categories(cats_uni)
        chart_uni.dLbls = DataLabelList(showVal=True)

        ws_sku.add_chart(chart_uni, f"{unique_chart_col}{pivot_start_uni}")

    # ======================= RIGHT SIDE: CUBIC ============================
    group_cub = cub_df[cub_df["Index_1"] == sku].dropna(subset=["Value"])
    if not group_cub.empty:
        pivot_cub = group_cub.pivot_table(index="Index_3", columns="Index_2",
                                          values="Value", aggfunc="sum", fill_value=0)
        pivot_cub.index.name = ""

        title_cell = f"{cubic_table_col}{pivot_start_uni - 1}"
        ws_sku[f"{get_column_letter(cubic_table_col)}{pivot_start_uni - 1}"] = f"{sku} - Cubic % (Forward + Reserve)"
        ws_sku[f"{get_column_letter(cubic_table_col)}{pivot_start_uni - 1}"].font = Font(size=14, bold=True)

        pivot_start_cub = pivot_start_uni

        pivot_cub_reset = pivot_cub.reset_index()
        for i in range(pivot_cub_reset.shape[0] + 1):
            for j in range(pivot_cub_reset.shape[1]):
                ws_sku.cell(row=pivot_start_cub + i,
                            column=cubic_table_col + j,
                            value=pivot_cub_reset.iloc[i - 1, j] if i > 0 else pivot_cub_reset.columns[j])

        pivot_end_cub = pivot_start_cub + len(pivot_cub)

        chart_cub = BarChart()
        chart_cub.type = "col"
        chart_cub.grouping = "stacked"
        chart_cub.overlap = 100
        chart_cub.title = f"{sku}: Cubic % Split (Forward + Reserve)"
        chart_cub.legend.position = "r"

        cats_cub = Reference(ws_sku, min_col=cubic_table_col, min_row=pivot_start_cub + 1,
                             max_row=pivot_end_cub )

        for col in range(cubic_table_col + 1, cubic_table_col + pivot_cub.shape[1] + 1):
            series = Series(
                Reference(ws_sku, min_row=pivot_start_cub + 1, max_row=pivot_end_cub ,
                         min_col=col, max_col=col),
                title=ws_sku.cell(row=pivot_start_cub, column=col).value
            )
            chart_cub.series.append(series)

        chart_cub.set_categories(cats_cub)
        chart_cub.dLbls = DataLabelList(showVal=True)

        ws_sku.add_chart(chart_cub, f"{cubic_chart_col}{pivot_start_cub}")

    row_pos = max(pivot_end_uni, pivot_end_cub) + 15

wb.save("Dashboard.xlsx")
print("Dashboard exported to Dashboard.xlsx")